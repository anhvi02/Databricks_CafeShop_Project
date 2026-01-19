"""
Create Balance Sheet Template Excel file.
This template defines the structure and hierarchy of balance sheet items.

Requires: pip install openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it using: pip install openpyxl")
    exit(1)

def create_balance_sheet_template():
    """Create Balance Sheet Template Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BS Template"
    
    # Define header style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define bold style for summary rows
    bold_font = Font(bold=True, size=11)
    
    # Define border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Row Index", "Balance Sheet Items", "Summary Items", "Balance Sheet Normalized", "Summary Index"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Define Balance Sheet structure for Coffee Shop
    # Format: (Row Index, Balance Sheet Items, Summary Items, Balance Sheet Normalized, Summary Index)
    # Summary Items and Summary Index are ONLY populated for summary/total rows
    # Indentation: Use spaces to create visual hierarchy (4 spaces per level for better readability)
    balance_sheet_items = [
        # Assets Section
        (1, "Assets", "", "Assets", ""),
        (2, "    Current Assets", "", "Current Assets", ""),
        (3, "        Cash", "", "Cash", ""),
        (4, "        Accounts receivable", "", "Accounts receivable", ""),
        (5, "        Inventory", "", "Inventory", ""),
        (6, "        Prepaid expenses", "", "Prepaid expenses", ""),
        (7, "        Short-term investments", "", "Short-term investments", ""),
        (8, "    Total current assets", "Total current assets", "Total current assets", 1),
        (9, "    Fixed (Long-Term) Assets", "", "Fixed (Long-Term) Assets", ""),
        (10, "        Long-term investments", "", "Long-term investments", ""),
        (11, "        Property, plant, and equipment", "", "Property, plant, and equipment", ""),
        (12, "        (Less accumulated depreciation)", "", "(Less accumulated depreciation)", ""),
        (13, "        Intangible assets", "", "Intangible assets", ""),
        (14, "    Total fixed assets", "Total fixed assets", "Total fixed assets", 2),
        (15, "    Other Assets", "", "Other Assets", ""),
        (16, "        Deferred income tax", "", "Deferred income tax", ""),
        (17, "        Other", "", "Other", ""),
        (18, "    Total Other Assets", "Total Other Assets", "Total Other Assets", 3),
        (19, "", "", "", ""),  # Empty row
        (20, "Total Assets", "Total Assets", "Total Assets", 4),
        (21, "", "", "", ""),  # Empty row
        
        # Liabilities and Owner's Equity Section
        (22, "Liabilities and Owner's Equity", "", "Liabilities and Owner's Equity", ""),
        (23, "    Current Liabilities", "", "Current Liabilities", ""),
        (24, "        Accounts payable", "", "Accounts payable", ""),
        (25, "        Short-term loans", "", "Short-term loans", ""),
        (26, "        Income taxes payable", "", "Income taxes payable", ""),
        (27, "        Accrued salaries and wages", "", "Accrued salaries and wages", ""),
        (28, "        Unearned revenue", "", "Unearned revenue", ""),
        (29, "        Current portion of long-term debt", "", "Current portion of long-term debt", ""),
        (30, "    Total current liabilities", "Total current liabilities", "Total current liabilities", 5),
        (31, "    Long-Term Liabilities", "", "Long-Term Liabilities", ""),
        (32, "        Long-term debt", "", "Long-term debt", ""),
        (33, "        Deferred income tax", "", "Deferred income tax", ""),
        (34, "        Other", "", "Other", ""),
        (35, "    Total long-term liabilities", "Total long-term liabilities", "Total long-term liabilities", 6),
        (36, "    Owner's Equity", "", "Owner's Equity", ""),
        (37, "        Owner's investment", "", "Owner's investment", ""),
        (38, "        Retained earnings", "", "Retained earnings", ""),
        (39, "        Other", "", "Other", ""),
        (40, "    Total owner's equity", "Total owner's equity", "Total owner's equity", 7),
        (41, "", "", "", ""),  # Empty row
        (42, "Total Liabilities and Owner's Equity", "Total Liabilities and Owner's Equity", "Total Liabilities and Owner's Equity", 8),
        (43, "", "", "", ""),  # Empty row
        
        # Common Financial Ratios Section
        (44, "Common Financial Ratios", "", "Common Financial Ratios", ""),
        (45, "    Debt Ratio (Total Liabilities / Total Assets)", "Debt Ratio (Total Liabilities / Total Assets)", "Debt Ratio (Total Liabilities / Total Assets)", 9),
        (46, "    Current Ratio (Current Assets / Current Liabilities)", "Current Ratio (Current Assets / Current Liabilities)", "Current Ratio (Current Assets / Current Liabilities)", 10),
        (47, "    Working Capital (Current Assets - Current Liabilities)", "Working Capital (Current Assets - Current Liabilities)", "Working Capital (Current Assets - Current Liabilities)", 11),
        (48, "    Assets-to-Equity Ratio (Total Assets / Owner's Equity)", "Assets-to-Equity Ratio (Total Assets / Owner's Equity)", "Assets-to-Equity Ratio (Total Assets / Owner's Equity)", 12),
        (49, "    Debt-to-Equity Ratio (Total Liabilities / Owner's Equity)", "Debt-to-Equity Ratio (Total Liabilities / Owner's Equity)", "Debt-to-Equity Ratio (Total Liabilities / Owner's Equity)", 13),
    ]
    
    # Populate data
    row_num = 2
    for item in balance_sheet_items:
        row_idx, item_name, summary, normalized, summary_idx = item
        
        # Row Index
        ws.cell(row=row_num, column=1, value=row_idx if row_idx else "")
        
        # Balance Sheet Items
        cell = ws.cell(row=row_num, column=2, value=item_name)
        if item_name and ("Total" in item_name or item_name.strip() in ["Assets", "Liabilities and Owner's Equity", "Common Financial Ratios"]):
            cell.font = bold_font
        
        # Summary Items - ONLY populate for summary rows (when summary is not empty)
        if summary:
            cell = ws.cell(row=row_num, column=3, value=summary)
            cell.font = bold_font
        else:
            ws.cell(row=row_num, column=3, value="")
        
        # Balance Sheet Normalized - mirrors Balance Sheet Items exactly (no indentation)
        cell = ws.cell(row=row_num, column=4, value=normalized)
        if normalized and ("Total" in normalized or normalized.strip() in ["Assets", "Liabilities and Owner's Equity", "Common Financial Ratios"]):
            cell.font = bold_font
        
        # Summary Index
        ws.cell(row=row_num, column=5, value=summary_idx if summary_idx else "")
        
        # Apply borders to all cells in row
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = thin_border
        
        # Alternate row colors (light blue and white)
        if row_num % 2 == 0:
            fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).fill = fill
        
        row_num += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    return wb


def main():
    """Main function to create the template."""
    print("Creating Balance Sheet Template...")
    
    wb = create_balance_sheet_template()
    
    output_file = '../data/financial/balance_sheet_template.xlsx'
    wb.save(output_file)
    
    print(f"Balance Sheet Template saved to {output_file}")
    print("\nTemplate includes:")
    print("  - Assets (Current, Fixed, Other)")
    print("  - Liabilities (Current, Long-Term)")
    print("  - Owner's Equity")
    print("  - Summary totals with indices")
    print("\nReady to use in Power BI for balance sheet reporting!")


if __name__ == '__main__':
    main()

