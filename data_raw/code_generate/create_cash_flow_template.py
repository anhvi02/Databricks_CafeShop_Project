"""
Create Cash Flow Template Excel file.
This template defines the structure and hierarchy of cash flow statement items.

Requires: pip install openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it using: pip install openpyxl")
    raise SystemExit(1)


def create_cash_flow_template():
    """Create Cash Flow Template Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CF Template"

    # Header style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Bold style for summary rows
    bold_font = Font(bold=True, size=11)

    # Border
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = ["Row Index", "Cash Flow Items", "Cash Flow Summary Items", "Cash Flow Normalized", "Summary Index"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Cash Flow structure for Coffee Shop
    # Format: (Row Index, Cash Flow Items, Cash Flow Summary Items, Cash Flow Normalized, Summary Index)
    # Summary Items and Summary Index are ONLY populated for summary/total rows
    # Indentation: 4 spaces per level for readability
    cash_flow_items = [
        # Operating Activities
        (1, "Operations", "", "Operations", ""),
        (2, "    Cash receipts from customers", "", "Cash receipts from customers", ""),
        (3, "    Cash paid for", "", "Cash paid for", ""),
        (4, "        Inventory purchases", "", "Inventory purchases", ""),
        (5, "        General operating expenses", "", "General operating and administrative expenses", ""),
        (6, "        Wage expenses", "", "Wage expenses", ""),
        (7, "    Interest", "", "Interest", ""),
        (8, "    Income taxes", "", "Income taxes", ""),
        (9, "Net Cash Flow from Operations", "Net Cash Flow from Operations", "Net Cash Flow from Operations", 1),
        (10, "", "", "", ""),

        # Investing Activities
        (11, "Investing Activities", "", "Investing Activities", ""),
        (12, "    Cash receipts from", "", "Cash receipts from", ""),
        (13, "        Sale of property and equipment", "", "Sale of property and equipment", ""),
        (14, "        Collection of principal on loans", "", "Collection of principal on loans", ""),
        (15, "        Sale of investment securities", "", "Sale of investment securities", ""),
        (16, "    Cash paid for", "", "Cash paid for", ""),
        (17, "        Purchase of property and equipment", "", "Purchase of property and equipment", ""),
        (18, "        Making loans to other entities", "", "Making loans to other entities", ""),
        (19, "        Purchase of investment securities", "", "Purchase of investment securities", ""),
        (20, "Net Cash Flow from Investing Activities", "Net Cash Flow from Investing Activities", "Net Cash Flow from Investing Activities", 2),
        (21, "", "", "", ""),

        # Financing Activities
        (22, "Financing Activities", "", "Financing Activities", ""),
        (23, "    Cash receipts from", "", "Cash receipts from", ""),
        (24, "        Issuance of stock", "", "Issuance of stock", ""),
        (25, "        Borrowing", "", "Borrowing", ""),
        (26, "    Cash paid for", "", "Cash paid for", ""),
        (27, "        Repurchase of stock (treasury stock)", "", "Repurchase of stock (treasury stock)", ""),
        (28, "        Repayment of loans", "", "Repayment of loans", ""),
        (29, "        Dividends", "", "Dividends", ""),
        (30, "Net Cash Flow from Financing Activities", "Net Cash Flow from Financing Activities", "Net Cash Flow from Financing Activities", 3),
        (31, "", "", "", ""),

        # Net Increase in Cash
        (32, "Net Increase in Cash", "Net Increase in Cash", "Net Increase in Cash", 4),
        (33, "", "", "", ""),

        # Cash at Beginning / End of Year
        (34, "Cash at Beginning of Year", "Cash at Beginning of Year", "Cash at Beginning of Year", 5),
        (35, "Cash at End of Year", "Cash at End of Year", "Cash at End of Year", 6),
    ]

    # Populate data
    row_num = 2
    for item in cash_flow_items:
        row_idx, item_name, summary, normalized, summary_idx = item

        # Row Index
        ws.cell(row=row_num, column=1, value=row_idx if row_idx else "")

        # Cash Flow Items
        cell = ws.cell(row=row_num, column=2, value=item_name)
        if item_name and (
            "Net Cash Flow" in item_name
            or item_name.strip() in ["Operations", "Investing Activities", "Financing Activities", "Net Increase in Cash"]
            or item_name.startswith("Cash at ")
        ):
            cell.font = bold_font

        # Cash Flow Summary Items - only for summary rows
        if summary:
            cell = ws.cell(row=row_num, column=3, value=summary)
            cell.font = bold_font
        else:
            ws.cell(row=row_num, column=3, value="")

        # Cash Flow Normalized
        cell = ws.cell(row=row_num, column=4, value=normalized)
        if normalized and (
            "Net Cash Flow" in normalized
            or normalized.strip() in ["Operations", "Investing Activities", "Financing Activities", "Net Increase in Cash"]
            or normalized.startswith("Cash at ")
        ):
            cell.font = bold_font

        # Summary Index
        ws.cell(row=row_num, column=5, value=summary_idx if summary_idx else "")

        # Borders
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = thin_border

        # Alternate row colors
        fill = PatternFill(
            start_color="D9E1F2" if row_num % 2 == 0 else "FFFFFF",
            end_color="D9E1F2" if row_num % 2 == 0 else "FFFFFF",
            fill_type="solid",
        )
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).fill = fill

        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 15

    # Freeze header
    ws.freeze_panes = "A2"

    return wb


def main():
    """Main function to create the Cash Flow template."""
    print("Creating Cash Flow Template...")
    wb = create_cash_flow_template()
    output_file = "../data/financial/cash_flow_template.xlsx"
    wb.save(output_file)
    print(f"Cash Flow Template saved to {output_file}")
    print("\nTemplate includes:")
    print("  - Operating Activities")
    print("  - Investing Activities")
    print("  - Financing Activities")
    print("  - Net Increase in Cash")
    print("  - Cash at Beginning / End of Year")
    print("\nReady to use in Power BI for cash flow reporting!")


if __name__ == "__main__":
    main()


