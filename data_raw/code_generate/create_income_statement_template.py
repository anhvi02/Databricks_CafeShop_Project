"""
Create Income Statement Template Excel file.
This template defines the structure and hierarchy of income statement items.

Requires: pip install openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it using: pip install openpyxl")
    exit(1)

def create_income_statement_template():
    """Create Income Statement Template Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "IS Template"
    
    # Define header style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define bold style for summary rows
    bold_font = Font(bold=True, size=11)
    
    # Define border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Row Index", "Income Statement Items", "Summary 1", "Items (Normalized)", "Summary Index"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Define Income Statement structure for Coffee Shop
    # Format: (Row Index, Income Statement Items, Summary 1, Items (Normalized), Summary Index)
    # Summary 1 and Summary Index are ONLY populated for summary/total rows
    # Indentation: Use spaces to create visual hierarchy (4 spaces per level for better readability)
    income_statement_items = [
        # Revenue Section
        (1, "    Total Sales Revenue", "", "Total Sales Revenue", ""),
        (2, "    Catering Revenue", "", "Catering Revenue", ""),
        (3, "    Delivery Revenue", "", "Delivery Revenue", ""),
        (4, "    Wholesale Revenue", "", "Wholesale Revenue", ""),
        (5, "Total Revenues", "Total Revenues", "Total Revenues", 1),
        (6, "", "", "", ""),  # Empty row
        
        # Cost of Goods Sold Section
        (7, "    Coffee Beans", "", "Coffee Beans", ""),
        (8, "    Milk & Dairy", "", "Milk & Dairy", ""),
        (9, "    Food Ingredients", "", "Food Ingredients", ""),
        (10, "    Packaging & Supplies", "", "Packaging & Supplies", ""),
        (11, "    Inventory Adjustments", "", "Inventory Adjustments", ""),
        (12, "Total COGS", "Total COGS", "Total COGS", 2),
        (13, "", "", "", ""),  # Empty row
        
        # Gross Profit Section
        (14, "Total Gross Profit", "Total Gross Profit", "Total Gross Profit", 3),
        (15, "Gross Profit %", "Gross Profit %", "Gross Profit %", 4),
        (16, "", "", "", ""),  # Empty row
        
        # Operating Expenses Section
        (17, "    Rent", "", "Rent", ""),
        (18, "    Utilities", "", "Utilities", ""),
        (19, "    Salaries & Wages", "", "Salaries & Wages", ""),
        (20, "    Marketing & Advertising", "", "Marketing & Advertising", ""),
        (21, "    Supplies", "", "Supplies", ""),
        (22, "    Equipment Maintenance", "", "Equipment Maintenance", ""),
        (23, "    Insurance", "", "Insurance", ""),
        (24, "    Professional Services", "", "Professional Services", ""),
        (25, "    Depreciation", "", "Depreciation", ""),
        (26, "    Other Operating Expenses", "", "Other Operating Expenses", ""),
        (27, "", "", "", ""),  # Empty row
        (28, "Total Other Expenses", "Total Other Expenses", "Total Other Expenses", 5),
        (29, "", "", "", ""),  # Empty row
        
        # Net Profit Section
        (30, "Total Net Profit", "Total Net Profit", "Total Net Profit", 6),
        (31, "Net Profit %", "Net Profit %", "Net Profit %", 7),
    ]
    
    # Populate data
    row_num = 2
    for item in income_statement_items:
        row_idx, item_name, summary, normalized, summary_idx = item
        
        # Row Index
        ws.cell(row=row_num, column=1, value=row_idx if row_idx else "")
        
        # Income Statement Items
        cell = ws.cell(row=row_num, column=2, value=item_name)
        if item_name and any(keyword in item_name for keyword in ["Total", "Profit", "%"]):
            cell.font = bold_font
        
        # Summary 1 - ONLY populate for summary rows (when summary is not empty)
        if summary:
            cell = ws.cell(row=row_num, column=3, value=summary)
            cell.font = bold_font
        else:
            ws.cell(row=row_num, column=3, value="")
        
        # Items (Normalized) - mirrors Income Statement Items exactly
        cell = ws.cell(row=row_num, column=4, value=normalized)
        if normalized and any(keyword in normalized for keyword in ["Total", "Profit", "%"]):
            cell.font = bold_font
        
        # Summary Index
        ws.cell(row=row_num, column=5, value=summary_idx if summary_idx else "")
        
        # Apply borders to all cells in row
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = thin_border
        
        # Alternate row colors (light blue and white)
        if row_num % 2 == 0:
            fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).fill = fill
        
        row_num += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    return wb


def main():
    """Main function to create the template."""
    print("Creating Income Statement Template...")
    
    wb = create_income_statement_template()
    
    output_file = '../data/financial/income_statement_template.xlsx'
    wb.save(output_file)
    
    print(f"Income Statement Template saved to {output_file}")
    print("\nTemplate includes:")
    print("  - Revenue items")
    print("  - Cost of Goods Sold items")
    print("  - Gross Profit calculations")
    print("  - Operating Expenses")
    print("  - Operating Profit")
    print("  - Other Income/Expenses")
    print("  - Net Profit")
    print("\nReady to use in Power BI for income statement reporting!")


if __name__ == '__main__':
    main()

